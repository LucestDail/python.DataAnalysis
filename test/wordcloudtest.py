from openpyxl import load_workbook

read_wb = load_workbook("./bludhouse.xlsx")
read_ws = read_wb.active

list_excel = []

for i in range(1,151) :
    print(read_ws.cell(i,1).value.strip())
    list_excel.append(str(read_ws.cell(i,1).value.strip()))

print(list_excel)

from konlpy.tag import Kkma
import collections
kkma = Kkma()



list_temp = []

for row in list_excel :
    list_temp = kkma.nouns(row) + list_temp
    
print(list_temp)

list_result = []
for check in list_temp :
    if len(check) > 1 :
        print(check)
        list_result.append(check)
        
last_text = ""

for i in list_result :
    last_text = last_text + " " + i
    
last_text

import matplotlib.pyplot as plt
from wordcloud import WordCloud


wordcloud = WordCloud(font_path="/Library/Fonts/NanumSquareOTFLight.otf").generate(last_text)

wordcloud.words_

plt.figure(figsize=(12,12))
plt.imshow(wordcloud, interpolation = "bilinear")
plt.axis("off")
plt.show()

from os import path
from PIL import Image
import numpy as np
import os

d = path.dirname(__file__) if "__file__" in locals() else os.getcwd()


mask = np.array(Image.open(path.join(d, 'cloud.png')))

last_text

wordcloud = WordCloud(
    font_path='/Library/Fonts/AppleGothic.ttf',
    background_color="white", 
    mask=mask,
).generate(last_text)
plt.figure(figsize=(12,12))
plt.imshow(wordcloud, interpolation='bilinear')
plt.axis("off")
plt.show()

# 이미지
# # https://pixabay.com/ko/vectors/%EA%B5%AC%EB%A6%84-%ED%95%98%EB%A3%A8-%EC%96%B4%EB%91%90%EC%9A%B4-cludy-37010/

# 코드 참조
# https://busy.org/@anpigon/2

# 한글이 안된다면?
# https://programmers.co.kr/learn/courses/21/lessons/950
# https://www.youtube.com/watch?v=XfLZH7-1pcM
# https://financedata.github.io/posts/matplotlib-hangul-for-windows-anaconda.html